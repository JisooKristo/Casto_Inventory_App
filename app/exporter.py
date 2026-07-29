from __future__ import annotations

from io import BytesIO

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter


EXPORT_COLUMNS = [
    "Item Name",
    "Company",
    "Location",
    "Device Type",
    "Serial Number",
    "Date Acquired",
    "Sequence Number",
    "Scan Timestamp",
]


def build_export_frame(records: list[dict[str, str]]) -> pd.DataFrame:
    normalized_records = []
    for record in records:
        normalized_records.append(
            {
                "Item Name": record.get("Item Name", "N/A"),
                "Company": record.get("Company", "N/A"),
                "Location": record.get("Location", "N/A"),
                "Device Type": record.get("Device Type", "N/A"),
                "Serial Number": record.get("Serial Number", record.get("serial_number", "N/A")),
                "Date Acquired": record.get("Date Acquired", "N/A"),
                "Sequence Number": record.get("Sequence Number", "N/A"),
                "Scan Timestamp": record.get("Scan Timestamp", "N/A"),
            }
        )

    return pd.DataFrame(normalized_records, columns=EXPORT_COLUMNS)


def generate_excel_bytes(records: list[dict[str, str]]) -> bytes:
    dataframe = build_export_frame(records)
    buffer = BytesIO()

    with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
        dataframe.to_excel(writer, index=False, sheet_name="Inventory Scans")

    buffer.seek(0)
    workbook = load_workbook(buffer)
    worksheet = workbook[workbook.sheetnames[0]]

    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)

    for cell in worksheet[1]:
        cell.fill = header_fill
        cell.font = header_font

    worksheet.freeze_panes = "A2"
    worksheet.auto_filter.ref = worksheet.dimensions

    for column_cells in worksheet.columns:
        cell_values = [str(cell.value) if cell.value is not None else "" for cell in column_cells]
        column_width = max(len(value) for value in cell_values) + 2 if cell_values else 12
        worksheet.column_dimensions[get_column_letter(column_cells[0].column)].width = min(column_width, 40)

    output = BytesIO()
    workbook.save(output)
    return output.getvalue()
